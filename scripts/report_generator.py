#!/usr/bin/env python3
"""
Amazon FBA Calculator Pro - Report Generator
Generates Excel reports from calculation data.
Can be run locally or as a simple API.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import json, sys

def generate_excel_report(products: list, output_path: str = "fba_report.xlsx"):
    """Generate a professional Excel FBA report from product list"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FBA利润分析"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="2E75B6")
    money_font = Font(bold=True, size=13)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["产品名", "售价($)", "成本($)", "头程($)", "ReferralFee($)", "FBA运费($)", "总费用($)", "净利润($)", "毛利率(%)", "ROI(%)", "结论"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    for row_idx, p in enumerate(products, 2):
        values = [
            p.get("name", f"产品{row_idx-1}"),
            p.get("price", 0),
            p.get("cost", 0),
            p.get("shipping", 0),
            p.get("referral_fee", 0),
            p.get("fulfillment_fee", 0),
            p.get("total_fees", 0),
            p.get("net_profit", 0),
            p.get("margin", 0),
            p.get("roi", 0),
            "✅盈利" if p.get("net_profit", 0) > 0 else "❌亏损"
        ]
        for col_idx, v in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            c.border = thin_border
            c.font = Font(size=11)
            if col_idx == 8:  # net profit
                c.font = Font(bold=True, size=12, color="10B981" if v > 0 else "EF4444")
            c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")

    # Column widths
    widths = [20, 10, 10, 8, 12, 12, 10, 10, 12, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 25
    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    # Demo: generate sample report
    sample = [
        {"name": "无线蓝牙耳机", "price": 29.99, "cost": 8.00, "shipping": 2.50, "referral_fee": 4.50, "fulfillment_fee": 4.55, "total_fees": 7.05, "net_profit": 12.44, "margin": 41.5, "roi": 155.5},
        {"name": "手机壳 iPhone15", "price": 19.99, "cost": 3.00, "shipping": 1.50, "referral_fee": 3.00, "fulfillment_fee": 3.73, "total_fees": 4.73, "net_profit": 10.76, "margin": 53.8, "roi": 358.7},
        {"name": "LED台灯", "price": 39.99, "cost": 15.00, "shipping": 4.00, "referral_fee": 6.00, "fulfillment_fee": 5.21, "total_fees": 11.21, "net_profit": 9.78, "margin": 24.5, "roi": 65.2},
    ]
    output = generate_excel_report(sample)
    print(f"Report generated: {output}")
